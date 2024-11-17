#Run a pokemon file
import numpy as np
import pandas as pd
import pokemon_module as pk
import csv
import openpyxl
from openpyxl import Workbook
import sys
import json


#Initial Conditions should be fed into python script
if len(sys.argv) >=2 :
    argstring = sys.argv[1]
    player_team = json.loads(argstring)
    #run index number
    #poke_index = int(sys.argv[1])
    #fight_index = int(sys.argv[2])
else:
    #Not enough inputs
    player_team = ['gyarados','zapdos','moltres','articuno','mew','mewtwo']
    #poke_index = 1
    #fight_index = 1

###Things to index: 
#1: which team we're on
#2: which battle we're on (changes per job index)

#read team from csv
#import csv
#with open('random_teams.csv', mode='r') as f:
#  reader = csv.reader(f)
#  #playerteam = reader(poke_index)
#  for current_line_number, row in enumerate(reader, start=1):
#        if current_line_number == poke_index:
#            player_team = row
#            break

# function to write multiple results to excel
def write_results_append(columns,results,sheet_name,file_name):
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()
        ws
    if sheet_name in wb.sheetnames:
        ws= wb[sheet_name]
    else:
        ws= wb.create_sheet(sheet_name)
        # Write the series index to the first row
        ws.append(columns)
        #for i, index_value in enumerate(results.index):
        #   ws.cell(row=1, column=i+1, value=index_value)
    
    string_list = [str(element) for element in results]

    ws.append(string_list)
    wb.save(file_name)
    return


#elite four teams
elite4_1 = ['dewgong','cloyster','slowbro','jynx','lapras']
elite4_2 = ['onix','hitmonlee','hitmonchan','onix','machamp']
elite4_3 = ['gengar','golbat','haunter','arbok','gengar']
elite4_4 = ['gyarados','dragonair','dragonair','aerodactyl','dragonite']
elite_list = [elite4_1,elite4_2,elite4_3,elite4_4]

#create pokemon objects
elite = []
for team in elite_list:
    elite.append(pk.create_pokemon_objects(team))

#player team
playerteam = pk.create_pokemon_objects(player_team)

#def sim_elite(team,elite,num_runs=10):
#    results = pd.DataFrame(columns=['Result','Time','Winner','Winner List'])
#    for i in range(num_runs):
#        result,time,teamname,winnerlist = pk.run_elite(team,elite,verbose=False,roundreset=False)
#        results.loc[i] = [result,time,teamname,winnerlist]
#    return results

columns = ['Result','Time','Winner','Winner List']
results = pk.run_elite(playerteam,elite,verbose=False)
#results = sim_elite(team,elite,num_runs=num_runs)

sheet_name = f"team{poke_index}"
file_name = "Output_data_files/elite_results_test.xlsx"


write_results_append(columns,results,sheet_name,file_name)
